import pandas as pd
import sqlite3
from datetime import datetime, timedelta
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelExporter:
    def __init__(self):
        self.db_path = 'instance/database.db'
    
    def get_db_connection(self):
        """Get database connection"""
        return sqlite3.connect(self.db_path)
    
    def get_tickets_data(self, start_date=None, end_date=None, status=None, category=None):
        """Get tickets data with optional filters"""
        conn = self.get_db_connection()
        
        query = """
        SELECT 
            t.id,
            t.title,
            t.description,
            t.category,
            t.priority,
            t.status,
            t.created_at,
            t.updated_at,
            u.username as assigned_agent,
            CASE 
                WHEN t.status = 'Resolved' THEN 
                    ROUND((julianday(t.updated_at) - julianday(t.created_at)) * 24, 2)
                ELSE NULL 
            END as resolution_time_hours
        FROM tickets t
        LEFT JOIN users u ON t.assigned_to = u.id
        WHERE 1=1
        """
        
        params = []
        
        if start_date:
            query += " AND DATE(t.created_at) >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND DATE(t.created_at) <= ?"
            params.append(end_date)
        
        if status:
            query += " AND t.status = ?"
            params.append(status)
        
        if category:
            query += " AND t.category = ?"
            params.append(category)
        
        query += " ORDER BY t.created_at DESC"
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        return df
    
    def get_analytics_summary(self):
        """Get analytics summary data"""
        conn = self.get_db_connection()
        
        # Overall statistics
        stats_query = """
        SELECT 
            COUNT(*) as total_tickets,
            COUNT(CASE WHEN status = 'Open' THEN 1 END) as open_tickets,
            COUNT(CASE WHEN status = 'In Progress' THEN 1 END) as in_progress_tickets,
            COUNT(CASE WHEN status = 'Resolved' THEN 1 END) as resolved_tickets,
            COUNT(CASE WHEN status = 'Closed' THEN 1 END) as closed_tickets,
            AVG(CASE WHEN status = 'Resolved' THEN 
                (julianday(updated_at) - julianday(created_at)) * 24 
            END) as avg_resolution_time_hours
        FROM tickets
        """
        
        stats_df = pd.read_sql_query(stats_query, conn)
        
        # Category breakdown
        category_query = """
        SELECT 
            category,
            COUNT(*) as ticket_count,
            COUNT(CASE WHEN status = 'Resolved' THEN 1 END) as resolved_count,
            AVG(CASE WHEN status = 'Resolved' THEN 
                (julianday(updated_at) - julianday(created_at)) * 24 
            END) as avg_resolution_time
        FROM tickets
        GROUP BY category
        ORDER BY ticket_count DESC
        """
        
        category_df = pd.read_sql_query(category_query, conn)
        
        # Priority breakdown
        priority_query = """
        SELECT 
            priority,
            COUNT(*) as ticket_count,
            COUNT(CASE WHEN status = 'Resolved' THEN 1 END) as resolved_count,
            AVG(CASE WHEN status = 'Resolved' THEN 
                (julianday(updated_at) - julianday(created_at)) * 24 
            END) as avg_resolution_time
        FROM tickets
        GROUP BY priority
        ORDER BY 
            CASE priority 
                WHEN 'Critical' THEN 1 
                WHEN 'High' THEN 2 
                WHEN 'Medium' THEN 3 
                WHEN 'Low' THEN 4 
            END
        """
        
        priority_df = pd.read_sql_query(priority_query, conn)
        
        # Daily ticket creation trend (last 30 days)
        trend_query = """
        SELECT 
            DATE(created_at) as date,
            COUNT(*) as tickets_created,
            COUNT(CASE WHEN status = 'Resolved' THEN 1 END) as tickets_resolved
        FROM tickets
        WHERE DATE(created_at) >= DATE('now', '-30 days')
        GROUP BY DATE(created_at)
        ORDER BY date
        """
        
        trend_df = pd.read_sql_query(trend_query, conn)
        
        conn.close()
        
        return {
            'overall_stats': stats_df,
            'category_breakdown': category_df,
            'priority_breakdown': priority_df,
            'daily_trend': trend_df
        }
    
    def style_worksheet(self, ws, title):
        """Apply professional styling to worksheet"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header styling to first row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add title
        ws.insert_rows(1)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:' + ws.max_column.__str__() + '1')
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def create_comprehensive_report(self, filters=None):
        """Create comprehensive Excel report with multiple sheets"""
        # Get data
        tickets_df = self.get_tickets_data(
            start_date=filters.get('start_date') if filters else None,
            end_date=filters.get('end_date') if filters else None,
            status=filters.get('status') if filters else None,
            category=filters.get('category') if filters else None
        )
        
        analytics_data = self.get_analytics_summary()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        summary_ws = wb.create_sheet("Executive Summary")
        
        # Add summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Tickets", analytics_data['overall_stats']['total_tickets'].iloc[0]],
            ["Open Tickets", analytics_data['overall_stats']['open_tickets'].iloc[0]],
            ["In Progress", analytics_data['overall_stats']['in_progress_tickets'].iloc[0]],
            ["Resolved Tickets", analytics_data['overall_stats']['resolved_tickets'].iloc[0]],
            ["Closed Tickets", analytics_data['overall_stats']['closed_tickets'].iloc[0]],
            ["Avg Resolution Time (Hours)", 
             round(analytics_data['overall_stats']['avg_resolution_time_hours'].iloc[0] or 0, 2)]
        ]
        
        for row in summary_data:
            summary_ws.append(row)
        
        self.style_worksheet(summary_ws, "Tech Support Dashboard - Executive Summary")
        
        # 2. All Tickets Sheet
        tickets_ws = wb.create_sheet("All Tickets")
        
        # Prepare tickets data for export
        export_tickets = tickets_df.copy()
        export_tickets['created_at'] = pd.to_datetime(export_tickets['created_at']).dt.strftime('%Y-%m-%d %H:%M')
        export_tickets['updated_at'] = pd.to_datetime(export_tickets['updated_at']).dt.strftime('%Y-%m-%d %H:%M')
        
        # Add data to worksheet
        for r in dataframe_to_rows(export_tickets, index=False, header=True):
            tickets_ws.append(r)
        
        self.style_worksheet(tickets_ws, "All Tickets Details")
        
        # 3. Category Analysis Sheet
        category_ws = wb.create_sheet("Category Analysis")
        
        category_data = analytics_data['category_breakdown'].copy()
        category_data['avg_resolution_time'] = category_data['avg_resolution_time'].round(2)
        
        for r in dataframe_to_rows(category_data, index=False, header=True):
            category_ws.append(r)
        
        self.style_worksheet(category_ws, "Tickets by Category Analysis")
        
        # 4. Priority Analysis Sheet
        priority_ws = wb.create_sheet("Priority Analysis")
        
        priority_data = analytics_data['priority_breakdown'].copy()
        priority_data['avg_resolution_time'] = priority_data['avg_resolution_time'].round(2)
        
        for r in dataframe_to_rows(priority_data, index=False, header=True):
            priority_ws.append(r)
        
        self.style_worksheet(priority_ws, "Tickets by Priority Analysis")
        
        # 5. Trend Analysis Sheet
        trend_ws = wb.create_sheet("Trend Analysis")
        
        trend_data = analytics_data['daily_trend'].copy()
        
        for r in dataframe_to_rows(trend_data, index=False, header=True):
            trend_ws.append(r)
        
        self.style_worksheet(trend_ws, "Daily Ticket Trends (Last 30 Days)")
        
        # Save to BytesIO for download
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def create_simple_export(self, data_type="tickets", filters=None):
        """Create simple Excel export for specific data type"""
        if data_type == "tickets":
            df = self.get_tickets_data(
                start_date=filters.get('start_date') if filters else None,
                end_date=filters.get('end_date') if filters else None,
                status=filters.get('status') if filters else None,
                category=filters.get('category') if filters else None
            )
            
            # Format dates
            df['created_at'] = pd.to_datetime(df['created_at']).dt.strftime('%Y-%m-%d %H:%M')
            df['updated_at'] = pd.to_datetime(df['updated_at']).dt.strftime('%Y-%m-%d %H:%M')
            
        elif data_type == "analytics":
            analytics_data = self.get_analytics_summary()
            df = analytics_data['category_breakdown']
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            
            # Style the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        output.seek(0)
        return output

# Global exporter instance
exporter = ExcelExporter()

def export_tickets_to_excel(filters=None):
    """Export tickets data to Excel"""
    return exporter.create_simple_export("tickets", filters)

def export_comprehensive_report(filters=None):
    """Export comprehensive analytics report to Excel"""
    return exporter.create_comprehensive_report(filters)

if __name__ == "__main__":
    # Test the exporter
    exporter = ExcelExporter()
    
    # Test comprehensive report
    output = exporter.create_comprehensive_report()
    
    # Save test file
    with open('test_report.xlsx', 'wb') as f:
        f.write(output.getvalue())
    
    print("Test report created: test_report.xlsx")
